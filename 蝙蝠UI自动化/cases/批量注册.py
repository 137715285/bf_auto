from appium import webdriver
from appium.webdriver.extensions.android.nativekey import AndroidKey
from openpyxl import load_workbook

desired_caps = {
  'platformName': 'Android', # 被测手机是安卓
  'platformVersion': '8', # 手机安卓版本
  'deviceName': 'xxx', # 设备名，安卓手机可以随意填写
  'appPackage': 'com.woyue.batchat', # 启动APP Package名称
  'appActivity': '.act.SplashActivity', # 启动Activity名称
  'unicodeKeyboard': True, # 使用自带输入法，输入中文时填True
  'resetKeyboard': True, # 执行完程序恢复原来输入法
  'noReset': True,       # 不要重置App
  'newCommandTimeout': 6000,
  'automationName' : 'UiAutomator2'
  # 'app': r'd:\apk\bili.apk',
}

# 连接Appium Server，初始化自动化环境
driver = webdriver.Remote('http://localhost:4723/wd/hub', desired_caps)

# 设置缺省等待时间
driver.implicitly_wait(5)

# 如果有`提示更新`界面，点击`×`
iknow = driver.find_element_by_id("ivClose")
if iknow:
    iknow.click()

# 根据id定位注册新账户，点击
driver.find_element_by_id("btnRegister").click()

# 根据id定位注册，点击
driver.find_element_by_id("tvOtherRegister").click()

# 读取excel
book = load_workbook(r'C:\Users\13771\Desktop\zc.xlsx')
sheet = book.active

vals =[]
for cell in sheet['A']:
    vals.append(cell.value)
print(vals[1:])

# 根据id定位昵称输入框
sbox = driver.find_element_by_id('et')
# 循环读取列表昵称，并逐个写入
for val in vals[1:]:
  uid_list = []
  sbox.send_keys(val)

  driver.find_element_by_id("btnNext").click()
  mm = driver.find_elements_by_id("et")
  mm[0].send_keys('1234567t')
  mm[1].send_keys('1234567t')
  driver.find_elements_by_id('btnNext').click()
  UID = driver.find_elements_by_id('tvBatId').text
  # 得到账号ID
  uid_list.append(UID)

#注册成功，确定按钮
#driver.find_elements_by_id('btnSure').click()
#返回上一页
#driver.pressKeyCode(AndroidKeyCode.BACK)
#写入excel
row = 2
for n in uid_list:
    sheet.cell(row, 2).value = n
    row += 1
book.save(r'C:\Users\13771\Desktop\zc.xlsx')